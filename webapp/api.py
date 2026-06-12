# -*- coding: utf-8 -*-
"""
Endpoints de la API REST para cálculo de vida útil.
"""
from __future__ import annotations

import csv
import io
from typing import List

from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse

from . import corrosion_model as cm
from .schemas import (
    CalcularVidaRequest,
    CalcularVidaResponse,
    CondicionCriticaResponse,
    DashboardCasoBase,
    DashboardComparativa,
    DashboardDescomposicion,
    DashboardHeatmap,
    DashboardKpis,
    DashboardRequest,
    DashboardResponse,
    DashboardSensibilidad,
    DashboardSerie,
    EscenariosResponse,
    EscenarioInfo,
    ExportRequest,
    MaterialesResponse,
    MaterialInfo,
    SensibilidadRequest,
    SensibilidadResponse,
    TablaRequest,
    TablaResponse,
)

router = APIRouter()


# =============================================================================
# METADATA
# =============================================================================

@router.get("/materiales", response_model=MaterialesResponse)
def get_materiales():
    """Devuelve materiales soportados y sus cédulas."""
    materiales = []
    for mat, data in cm.MATERIALS.items():
        desc = {
            'SS304': 'Austenítico inoxidable SS304 (sin soldadura, E=1.0)',
            'SS304L': 'SS304L dual certificado TP304/304L (S del TP304, γ=0.85)',
            'A53 GRB': 'Acero al carbono A53 Gr B ERW (E=0.85)',
        }[mat]
        materiales.append(MaterialInfo(material=mat, schedules=data['schedules'], descripcion=desc))
    return MaterialesResponse(materiales=materiales, diametros=cm.DIAMETROS_ORDEN)


@router.get("/escenarios", response_model=EscenariosResponse)
def get_escenarios():
    """Devuelve escenarios de servicio con parámetros típicos."""
    escenarios = []
    for key, esc in cm.ESCENARIOS.items():
        escenarios.append(EscenarioInfo(
            id=key,
            nombre=esc['nombre'],
            T_ref=esc['T_ref'],
            v_tip=esc['v_tip'],
            Cs_tip=esc['Cs_tip'],
            W_quim_SS=esc['W_quim_SS'],
            W_quim_A53=esc['W_quim_A53'],
        ))
    return EscenariosResponse(escenarios=escenarios)


# =============================================================================
# CÁLCULOS
# =============================================================================

@router.post("/calcular/vida-util", response_model=CalcularVidaResponse)
def calcular_vida_util(req: CalcularVidaRequest):
    """Calcula la vida útil para una tubería, material y condiciones dadas."""
    try:
        OD = cm.get_OD(req.material, req.nps)
        t_nom = cm.get_t_nom(req.material, req.nps, req.schedule)
        res = cm.calc_vida(
            OD=OD,
            t_nom=t_nom,
            material=req.material,
            escenario=req.escenario,
            Cs=req.Cs,
            v=req.v,
            T=req.T,
            beta=req.beta,
            k0_override=req.k0_override,
            fs=req.fs,
        )
        T_eff = req.T if req.T is not None else cm.ESCENARIOS[req.escenario.capitalize()]['T_ref']
        return CalcularVidaResponse(
            material=req.material,
            nps=req.nps,
            schedule=req.schedule,
            escenario=req.escenario.capitalize(),
            Cs=req.Cs,
            v=req.v,
            T=T_eff,
            fs=req.fs,
            vida_util=round(res['vida_util'], 2),
            t_disp=round(res['t_disp'], 3),
            t_min=round(res['t_min'], 3),
            W_quim=round(res['W_quim'], 5),
            W_ero=round(res['W_ero'], 5),
            W_total=round(res['W_total'], 5),
            S=round(res['S'], 1),
            unidades={
                'vida_util': 'años',
                't_disp': 'mm',
                't_min': 'mm',
                'W_quim': 'mm/año',
                'W_ero': 'mm/año',
                'W_total': 'mm/año',
                'S': 'MPa',
            },
            advertencias=res['advertencias'],
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {e}")


@router.post("/calcular/tabla", response_model=TablaResponse)
def calcular_tabla(req: TablaRequest):
    """Genera tabla paramétrica de vida útil por NPS y cédula."""
    try:
        resultados = cm.calcular_tabla(
            material=req.material,
            Cs=req.Cs,
            v=req.v,
            escenario=req.escenario,
            T=req.T,
            beta=req.beta,
            k0_override=req.k0_override,
            fs=req.fs,
        )
        esc_for_adv = req.escenario if req.escenario else 'Alcalino'
        advertencias = cm.generar_advertencias(
            req.material, esc_for_adv, req.Cs, req.v, req.T, req.beta, req.k0_override, req.fs
        )
        return TablaResponse(
            material=req.material,
            Cs=req.Cs,
            v=req.v,
            T=req.T,
            fs=req.fs,
            beta=req.beta,
            escenarios=list(resultados.keys()),
            diametros=cm.DIAMETROS_ORDEN,
            schedules=cm.MATERIALS[req.material]['schedules'],
            valores=resultados,
            advertencias=advertencias,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {e}")


@router.post("/calcular/sensibilidad-fs", response_model=SensibilidadResponse)
def calcular_sensibilidad(req: SensibilidadRequest):
    """Calcula vida útil para distintos factores de seguridad."""
    try:
        OD = cm.get_OD(req.material, req.nps)
        t_nom = cm.get_t_nom(req.material, req.nps, req.schedule)
        resultados = cm.sensibilidad_fs(
            OD=OD,
            t_nom=t_nom,
            material=req.material,
            escenario=req.escenario,
            Cs=req.Cs,
            v=req.v,
            T=req.T,
            beta=req.beta,
            k0_override=req.k0_override,
            fs_vals=tuple(req.fs_vals),
        )
        advertencias = cm.generar_advertencias(
            req.material, req.escenario, req.Cs, req.v, req.T, req.beta, req.k0_override, req.fs_vals[0]
        )
        return SensibilidadResponse(
            material=req.material,
            nps=req.nps,
            schedule=req.schedule,
            escenario=req.escenario.capitalize(),
            Cs=req.Cs,
            v=req.v,
            T=req.T,
            resultados={str(k): round(v, 2) for k, v in resultados.items()},
            advertencias=advertencias,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {e}")


@router.get("/calcular/condicion-critica", response_model=CondicionCriticaResponse)
def calcular_condicion_critica():
    """Devuelve la tabla de condición crítica del Resumen Ejecutivo."""
    filas = cm.condicion_critica_60C()
    advertencias = [
        "Condición crítica calibrada a 60 °C con pulpa refinada (β = 0,25, k₀,ref = 0,0035).",
        "Velocidad de 3,5 m/s excede el rango donde se desprecia sinergia erosión-corrosión (Ws ≈ 0).",
        "Requiere validación con mediciones UT de planta.",
    ]
    return CondicionCriticaResponse(
        parametros={
            'T': 60.0,
            'Cs': 6.0,
            'v': 3.5,
            'beta': 0.25,
            'k0_ref': 0.0035,
            'fs': 1.1,
        },
        filas=filas,
        advertencias=advertencias,
    )


# =============================================================================
# DASHBOARD EJECUTIVO (derivado de los inputs de la calculadora)
# =============================================================================

def _sch_equivalente(material: str, schedule: str) -> str:
    """Mapea la cédula entre familias (10S <-> 10) para comparar materiales."""
    if material == 'A53 GRB':
        return schedule.replace('S', '')
    return schedule if schedule.endswith('S') else schedule + 'S'


def _k0_para_material(material: str, k0_override) -> float | None:
    """
    Si el usuario aplica k0 reducido por fibra refinada (propiedad del fluido),
    se escala simétricamente al A53 (trato simétrico, decisión C-4b del informe).
    """
    if k0_override is None:
        return None
    if material == 'A53 GRB':
        return cm.MATERIALS['A53 GRB']['k0'] * (k0_override / cm.MATERIALS['SS304']['k0'])
    return k0_override


@router.post("/dashboard", response_model=DashboardResponse)
def post_dashboard(req: DashboardRequest):
    """
    Resumen Ejecutivo calculado con los datos vivos de la calculadora:
    KPIs por material, heatmap por escenario, comparativa por NPS,
    descomposición de tasas y sensibilidad al FS — todo a las condiciones dadas.
    """
    try:
        materiales = list(cm.MATERIALS.keys())
        escenarios = list(cm.ESCENARIOS.keys())
        T_eff = req.T if req.T is not None else cm.ESCENARIOS[req.escenario]['T_ref']

        # --- KPIs: los 3 materiales al NPS y cédula equivalente seleccionados
        detalle = {}
        for mat in materiales:
            sch = _sch_equivalente(mat, req.schedule)
            detalle[mat] = cm.calc_vida(
                OD=cm.get_OD(mat, req.nps),
                t_nom=cm.get_t_nom(mat, req.nps, sch),
                material=mat,
                escenario=req.escenario,
                Cs=req.Cs,
                v=req.v,
                T=req.T,
                beta=req.beta,
                k0_override=_k0_para_material(mat, req.k0_override),
                fs=req.fs,
            )
        sel = detalle[req.material]
        kpis = DashboardKpis(
            vida_ss304=round(detalle['SS304']['vida_util'], 1),
            vida_ss304l=round(detalle['SS304L']['vida_util'], 1),
            vida_a53=round(detalle['A53 GRB']['vida_util'], 1),
            t_disp=round(sel['t_disp'], 2),
            W_quim=round(sel['W_quim'], 4),
            W_ero=round(sel['W_ero'], 4),
            W_total=round(sel['W_total'], 4),
            S=round(sel['S'], 1),
            unidades={'vida': 'años', 't_disp': 'mm', 'W': 'mm/año', 'S': 'MPa'},
        )
        caso_base = DashboardCasoBase(
            descripcion=(
                f"{cm.ESCENARIOS[req.escenario]['nombre']} · T = {T_eff:.0f} °C · "
                f"v = {req.v:.1f} m/s · Cs = {req.Cs:.1f} % · FS = {req.fs:.1f}"
                + (f" · pulpa refinada (β = {req.beta:.2f})" if req.beta > 0 else "")
            ),
            parametros={'T': T_eff, 'Cs': req.Cs, 'v': req.v, 'beta': req.beta,
                        'k0_override': req.k0_override, 'fs': req.fs},
            material_ref=req.material,
            nps_ref=req.nps,
            schedule_ref=req.schedule,
            escenario_ref=req.escenario,
            kpis=kpis,
        )

        # --- Heatmap material × escenario a las condiciones del usuario
        valores = []
        for mat in materiales:
            sch = _sch_equivalente(mat, req.schedule)
            OD = cm.get_OD(mat, req.nps)
            t_nom = cm.get_t_nom(mat, req.nps, sch)
            fila = []
            for e in escenarios:
                res = cm.calc_vida(OD, t_nom, mat, e, Cs=req.Cs, v=req.v, T=req.T,
                                   beta=req.beta, k0_override=_k0_para_material(mat, req.k0_override), fs=req.fs)
                fila.append(round(res['vida_util'], 1))
            valores.append(fila)
        heatmap = DashboardHeatmap(
            materiales=materiales,
            escenarios=escenarios,
            escenarios_nombres=[cm.ESCENARIOS[e]['nombre'] for e in escenarios],
            valores=valores,
            nota=f"{req.nps} Sch {req.schedule} equivalente · T = {T_eff:.0f} °C · v = {req.v:.1f} m/s · Cs = {req.Cs:.1f} %",
        )

        # --- Comparativa por NPS: 10S y 40S de inoxidables + A53 Sch 40
        series_def = [
            ('ss304_s10', 'SS304 Sch 10S', 'SS304', '10S'),
            ('ss304L_s10', 'SS304L Sch 10S', 'SS304L', '10S'),
            ('ss304_s40', 'SS304 Sch 40S', 'SS304', '40S'),
            ('ss304L_s40', 'SS304L Sch 40S', 'SS304L', '40S'),
            ('a53_s40', 'A53 Gr B Sch 40', 'A53 GRB', '40'),
        ]
        series = []
        for sid, nombre, mat, sch in series_def:
            vals = []
            for nps in cm.DIAMETROS_ORDEN:
                res = cm.calc_vida(
                    cm.get_OD(mat, nps), cm.get_t_nom(mat, nps, sch), mat, req.escenario,
                    Cs=req.Cs, v=req.v, T=req.T, beta=req.beta,
                    k0_override=_k0_para_material(mat, req.k0_override), fs=req.fs,
                )
                vals.append(round(res['vida_util'], 1))
            series.append(DashboardSerie(id=sid, nombre=nombre, valores=vals))
        comparativa = DashboardComparativa(
            nps=[n.replace('NPS ', '') for n in cm.DIAMETROS_ORDEN],
            series=series,
        )

        # --- Descomposición W_quim vs W_ero por material (NPS/cédula seleccionados)
        descomposicion = DashboardDescomposicion(
            materiales=materiales,
            W_quim=[round(detalle[m]['W_quim'], 4) for m in materiales],
            W_ero=[round(detalle[m]['W_ero'], 4) for m in materiales],
            unidad='mm/año',
        )

        # --- Sensibilidad al FS para el material/NPS/cédula seleccionados
        fs_vals = tuple(round(1.0 + 0.1 * i, 1) for i in range(11))  # 1.0 … 2.0
        sens = cm.sensibilidad_fs(
            OD=cm.get_OD(req.material, req.nps),
            t_nom=cm.get_t_nom(req.material, req.nps, req.schedule),
            material=req.material,
            escenario=req.escenario,
            Cs=req.Cs,
            v=req.v,
            T=req.T,
            beta=req.beta,
            k0_override=req.k0_override,
            fs_vals=fs_vals,
        )
        sensibilidad = DashboardSensibilidad(
            material=req.material,
            nps=req.nps,
            schedule=req.schedule,
            fs=list(fs_vals),
            vida=[round(sens[f], 1) for f in fs_vals],
        )

        advertencias = list(sel['advertencias'])
        advertencias.append(
            "SS304L modelado como dual certificado TP304/TP304L (ASTM A312): diseño con S del TP304 "
            "y química de bajo carbono; especificarlo así en la orden de compra."
        )
        if req.k0_override is not None:
            advertencias.append(
                "El coeficiente de erosión reducido por fibra refinada se aplica simétricamente al A53 "
                "(propiedad del fluido, no del material)."
            )

        return DashboardResponse(
            caso_base=caso_base,
            heatmap=heatmap,
            comparativa=comparativa,
            descomposicion=descomposicion,
            sensibilidad_fs=sensibilidad,
            advertencias=advertencias,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {e}")


# =============================================================================
# EXPORTACIÓN
# =============================================================================

def _flatten_tabla(req: ExportRequest) -> List[List[str]]:
    """Aplana la tabla paramétrica para exportación."""
    resultados = cm.calcular_tabla(
        material=req.material,
        Cs=req.Cs,
        v=req.v,
        escenario=req.escenario,
        T=req.T,
        beta=req.beta,
        k0_override=req.k0_override,
        fs=req.fs,
    )
    schedules = cm.MATERIALS[req.material]['schedules']
    rows: List[List[str]] = []
    # Header
    header = ['Escenario', 'NPS'] + schedules
    rows.append(header)
    for esc, nps_data in resultados.items():
        for nps in cm.DIAMETROS_ORDEN:
            row = [esc, nps]
            for sch in schedules:
                row.append(str(cm.format_vida(nps_data[nps][sch])))
            rows.append(row)
    return rows


@router.post("/export/csv")
def export_csv(req: ExportRequest):
    """Exporta la tabla paramétrica a CSV."""
    try:
        rows = _flatten_tabla(req)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        output.seek(0)
        filename = f"vida_util_{req.material.replace(' ', '_')}_Cs{req.Cs}_v{req.v}.csv"
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {e}")


@router.post("/export/excel")
def export_excel(req: ExportRequest):
    """Exporta la tabla paramétrica a Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        rows = _flatten_tabla(req)
        wb = Workbook()
        ws = wb.active
        ws.title = "Vida útil"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        for c in range(3, 3 + len(rows[0])):
            ws.column_dimensions[chr(64 + c) if c <= 26 else chr(64 + c - 26)].width = 12

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"vida_util_{req.material.replace(' ', '_')}_Cs{req.Cs}_v{req.v}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está disponible en el entorno")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {e}")


# =============================================================================
# DESCARGAS DE ENTREGABLES
# =============================================================================

DOWNLOADS_DIR = Path(__file__).resolve().parent / "static" / "downloads"


@router.get("/descargas/informe")
def descargar_informe():
    """Descarga el informe técnico P2603-PR-INF-001 Rev.2 (PDF)."""
    f = DOWNLOADS_DIR / "P2603-PR-INF-001_Rev2.pdf"
    if not f.exists():
        raise HTTPException(
            status_code=404,
            detail="El informe PDF no está disponible en el servidor. Copie el archivo a webapp/static/downloads/.",
        )
    return FileResponse(f, media_type="application/pdf", filename="P2603-PR-INF-001_Rev2.pdf")


@router.get("/descargas/presentacion")
def descargar_presentacion():
    """Descarga la presentación P2603-PR-PPT-001 Rev.2 (PPTX)."""
    f = DOWNLOADS_DIR / "P2603-PR-PPT-001_Rev2.pptx"
    if not f.exists():
        raise HTTPException(
            status_code=404,
            detail="La presentación PPTX no está disponible en el servidor. Copie el archivo a webapp/static/downloads/.",
        )
    return FileResponse(
        f,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        filename="P2603-PR-PPT-001_Rev2.pptx",
    )
